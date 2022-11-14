import datetime
import socket
from time import sleep

import nidmm
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import *
from openpyxl import load_workbook
from openpyxl.styles import Side, Border

from Reports_Generator import Get_Reports
from Ui_Browse_Test import Ui_Dialog_BrowseTest
from Dlg_FaultReport import FaultReportDlg
import os
####################################################################################################################
class BrowseTestDlg(QDialog,Ui_Dialog_BrowseTest):
    def __init__(self,parent = None):
        super().__init__(parent)
        self.setupUi(self)
        self.setWindowFlags(Qt.WindowMinimizeButtonHint)
        self.AbortTestFlag = False
        self.TestFailFlag = False
        #self.min = 0
        #self.max = 5
        self.FaultReportDlg = FaultReportDlg()
        self.tableWidget.setRowCount(128)
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.pushButton_Back.clicked.connect(self.closewindow)

        self.pushButton_Measure.clicked.connect(self.fun_measure)
        self.pushButton_Abort.clicked.connect(self.AbortTest)
        self.pushButton_Save.clicked.connect(self.SaveReport)
        self.pushButton_Abort.setDisabled(True)
        self.pushButton_Save.setDisabled(True)
        self.pushButton_Measure.setDisabled(True)


        self.progressBar.setValue(0)

        self.pushButton_Browse.clicked.connect(self.fun_browse)

        self.progressBar.setValue(0)

        self.frompinlist = []
        self.topinlist = []
        self.minvallist = []
        self.maxvallist = []

    def closewindow(self):
        self.close()
    #################################################################################################################
    def AbortTest(self):
        msg = QMessageBox.critical(self,"Browse Test","Do you want to Abort?",QMessageBox.Yes|QMessageBox.No)
        if msg==QMessageBox.Yes:
            self.AbortTestFlag = True
            self.tableWidget.setRowCount(self.tableWidget.currentRow()+1)
            '''
            if self.TestFailFlag == True:
                frmsg = QMessageBox.question(self, "Failure Report", "Do you want to view the Failure Report",
                                             QMessageBox.Yes | QMessageBox.No)
                self.TestFailFlag=False
                if frmsg == QMessageBox.Yes:
                    self.TestFailFlag = False
                    dlg = self.FaultReportDlg
                    dlg.exec()
            '''
            # This line removed by Bhaskar Rao on 10 Nov 2022
            #self.tableWidget.setRowCount(self.tableWidget.currentRow()+1)
            self.pushButton_Back.setEnabled(True)
            self.pushButton_Save.setEnabled(True)
            self.pushButton_Measure.setEnabled(True)
        else:
            self.AbortTestFlag = False
            self.pushButton_Back.setDisabled(True)
            self.pushButton_Save.setDisabled(True)
            self.pushButton_Measure.setDisabled(True)
    ###################################################################################################################
    def fun_measure(self):

        try:
            sock = socket.socket()
            sock.connect(('192.168.1.10', 5003))
        except:
            print('unable to connect to server')
            QMessageBox.information(self, "Communication Link Down", "Unable to Communicate with  Interface Box")
            return
        try:
            session = nidmm.Session("DMM4605")
            session.configure_measurement_digits(measurement_function=nidmm.Function["TWO_WIRE_RES"], range=10e3,
                                                 resolution_digits=6.5)
        except:
            QMessageBox.information(self, "Communication Link Down", "Unable to Communicate with  DMM")
            return

        self.pushButton_Abort.setEnabled(True)
        self.pushButton_Back.setDisabled(True)
        self.pushButton_Save.setDisabled(True)
        self.pushButton_Measure.setDisabled(True)
        self.tableWidget.setRowCount(128)
        self.tableWidget.clear()
        self.tableWidget.setHorizontalHeaderLabels(
            ["S.No", "Line X", "Linw Y", "Min", "Measured ", "Max", "Units", "Result"])
        self.AbortTestFlag = False
        measured_value = 4.5
        FailTrailCount = 0
        falutypoints = 0
        try:
            frominlist_byte = bytearray(self.frompinlist)
            topinlist_byte = bytearray(self.topinlist)
        except:
            QMessageBox.warning(self,"Error","Invalid File Format")

        #print(type(frominlist_byte[1]))
        for i in range(0,self.nr):
            self.tableWidget.setItem(i, 4, QTableWidgetItem('---'))
            self.tableWidget.setItem(i, 7, QTableWidgetItem("---"))
        i = 0
        while (self.AbortTestFlag == False) and (i < self.nr-8):
            Packet = []
            Packet.append(0xCC)
            Packet.append(frominlist_byte[i])
            print('Lane_X',frominlist_byte[i])
            Packet.append(0x01)
            Packet.append(topinlist_byte[i])
            print('Lane_Y', topinlist_byte[i])
            Packet.append(0x01)
            try:
                sock.send(bytes(Packet))
                sleep(0.1)
            except:
                print('Tansmission Failed')
                # QMessageBox.information(self, "Communication Link Down", "Unable to Communicate with  Hardware")
                msg = QMessageBox.critical(self, "Link Down", "Unable to Communicate with Interface Box\nDo you want to Retry?",
                                           QMessageBox.Yes | QMessageBox.No)
                if msg == QMessageBox.Yes:
                    QMessageBox.information(self, "Link Down",
                                            "Restart Interface Box\nWait till LAN LEDs Blinking on front Panel")
                    sock.close()
                    try:
                        sock = socket.socket()
                        sock.connect(('192.168.1.10', 5003))
                        #ComErrFlag = True
                    except:
                        print('unable to connect to server')
                        QMessageBox.information(self, "Communication Link Down",
                                                "Unable to Communicate with  Interface Box")
                        return
                else:
                    self.AbortTestFlag = True
                    sock.close()


            self.tableWidget.setItem(i, 0, QTableWidgetItem(str(i + 1)))
            self.tableWidget.setItem(i, 1, QTableWidgetItem("PIN-" + str(frominlist_byte[i])))
            self.tableWidget.setItem(i, 2, QTableWidgetItem("PIN-" + str(topinlist_byte[i])))
            #self.tableWidget.setItem(i, 3, QTableWidgetItem(str(self.min)))

            #self.tableWidget.setItem(i, 5, QTableWidgetItem(str(self.max)))
            self.tableWidget.setItem(i, 6, QTableWidgetItem("Ohms"))

            self.tableWidget.setRowCount(self.nr)
            self.tableWidget.selectRow(i)
            self.progressBar.setValue(int((i + 1) * 100 / self.nr))

            qApp.processEvents()
            measured_value = self.GetMeasfromDMM(session=session, range=10e3)
            if measured_value == None:
                msg = QMessageBox.critical(self, "Link Down", "Unable to Communicate with  DMM\nDo you want to Retry?",
                                           QMessageBox.Yes | QMessageBox.No)
                if msg == QMessageBox.Yes:
                    QMessageBox.information(self, "Link Down",
                                            "Unplug USB Cable of DMM & re-plug. wait for few seconds")
                    try:
                        session = nidmm.Session("DMM4605")
                        session.configure_measurement_digits(measurement_function=nidmm.Function["TWO_WIRE_RES"],
                                                             range=10e3,
                                                             resolution_digits=6.5)
                        ComErrFlag = True
                    except:
                        #QMessageBox.information(self, "Communication Link Down", "Unable to Communicate with  DMM")
                        return
                else:
                    self.AbortTestFlag = True
                    self.tableWidget.setRowCount(self.tableWidget.currentRow()+1)
                '''
                if FailTrailCount >= 3:
                    self.tableWidget.setItem(i, 4, QTableWidgetItem(str('>100K')))
                    self.tableWidget.setItem(i, 7, QTableWidgetItem("FAILED"))
                    FailTrailCount = FailTrailCount + 1
                '''
            else:
                self.tableWidget.setItem(i, 4, QTableWidgetItem(f'''{measured_value:.2f}'''))
                self.min = self.minvallist[i]
                self.max = self.maxvallist[i]
                self.tableWidget.setItem(i, 3, QTableWidgetItem(f'''{self.min:.2f}'''))
                self.tableWidget.setItem(i, 5, QTableWidgetItem(f'''{self.max:.2f}'''))
                print(self.min,self.max)
                if measured_value > self.min and measured_value < self.max:
                    self.tableWidget.setItem(i, 7, QTableWidgetItem("PASS"))
                    i = i + 1
                    FailTrailCount = 0
                elif FailTrailCount >= 3:
                    self.tableWidget.setItem(i, 7, QTableWidgetItem("FAILED"))
                    self.FaultReportDlg.tableWidget.setRowCount(falutypoints + 1)
                    self.FaultReportDlg.label.setText("BROWSE TEST-FAULTS REPORT")
                    self.FaultReportDlg.GUI="BrowseTest"

                    self.FaultReportDlg.tableWidget.setItem(falutypoints, 0,
                                                            QTableWidgetItem(self.tableWidget.item(i, 0).text()))

                    self.FaultReportDlg.tableWidget.setItem(falutypoints, 1,
                                                            QTableWidgetItem(self.tableWidget.item(i, 1).text()))
                    self.FaultReportDlg.tableWidget.setItem(falutypoints, 2,
                                                            QTableWidgetItem(self.tableWidget.item(i, 2).text()))
                    self.FaultReportDlg.tableWidget.setItem(falutypoints, 3,
                                                            QTableWidgetItem(self.tableWidget.item(i, 3).text()))
                    self.FaultReportDlg.tableWidget.setItem(falutypoints, 4,
                                                            QTableWidgetItem(self.tableWidget.item(i, 4).text()))
                    self.FaultReportDlg.tableWidget.setItem(falutypoints, 5,
                                                            QTableWidgetItem(self.tableWidget.item(i, 5).text()))
                    self.FaultReportDlg.tableWidget.setItem(falutypoints, 6,
                                                            QTableWidgetItem(self.tableWidget.item(i, 6).text()))
                    self.FaultReportDlg.tableWidget.setItem(falutypoints, 7,
                                                            QTableWidgetItem(self.tableWidget.item(i, 7).text()))
                    # self.FaultReportDlg.tableWidget.setItem(falutypoints, 0, self.tableWidget.item(i, 0).text())

                    falutypoints = falutypoints + 1
                    self.TestFailFlag=True
                    i = i + 1
                    FailTrailCount = 0
                else:
                    FailTrailCount = FailTrailCount + 1
        sock.close()
        self.pushButton_Back.setEnabled(True)
        self.pushButton_Abort.setDisabled(True)
        self.pushButton_Save.setEnabled(True)
        print("measure")
        if self.TestFailFlag==True:
            self.TestFailFlag=False
            frmsg = QMessageBox.question(self,"Failure Report","Do you want to view the Failure Report",QMessageBox.Yes|QMessageBox.No)
            if frmsg==QMessageBox.Yes:
                dlg = self.FaultReportDlg
                dlg.exec()
    ##############################################################################################################
    def GetMeasfromDMM(self,session =None,range = 100e6):
        #with nidmm.Session("DMM4605") as session:

        try:
            meas_res = session.read()
            return meas_res
        except:
            print("out of range")
            #QMessageBox.information(self, "Communication Link Down", "Unable to Communicate with  DMM")
            return None
    #################################################################################################################
    def fun_browse(self):
        path = QFileDialog.getOpenFileName(self, "Select File", os.getenv("Home"))
        if len(path[0]) > 0:
            try:
                self.lineEdit.setText(path[0])
                self.frompinlist = []
                self.topinlist = []
                self.minvallist = []
                self.maxvallist = []
                workbook = load_workbook(filename=path[0])
                self.inputsheet = workbook.active
                self.nr = self.inputsheet.max_row
                for i in range(8, self.nr):
                    self.frompinlist.append(self.inputsheet.cell(i, 1).value)
                    self.topinlist.append(self.inputsheet.cell(i,3).value)
                    self.minvallist.append(self.inputsheet.cell(i,6).value)
                    self.maxvallist.append(self.inputsheet.cell(i, 7).value)
                self.tableWidget.setRowCount(self.nr-8)
                for i in range(0,self.nr-8):
                    self.tableWidget.setItem(i, 0, QTableWidgetItem(str(i + 1)))
                    self.tableWidget.setItem(i, 1, QTableWidgetItem("PIN-" + str(self.frompinlist[i])))
                    self.tableWidget.setItem(i, 2, QTableWidgetItem("PIN-" + str(self.topinlist[i])))
                    self.tableWidget.setItem(i, 3, QTableWidgetItem(str(self.minvallist[i])))
                    self.tableWidget.setItem(i, 4, QTableWidgetItem("---"))
                    self.tableWidget.setItem(i, 5, QTableWidgetItem(str(self.maxvallist[i])))
                    self.tableWidget.setItem(i, 6, QTableWidgetItem("Ohms"))
                    self.tableWidget.setItem(i, 7, QTableWidgetItem("---"))
                    self.pushButton_Measure.setDisabled(False)

            except:
                self.lineEdit.clear()
                self.tableWidget.clear()
                self.tableWidget.setRowCount(0)
                self.tableWidget.setHorizontalHeaderLabels(["S.No","Line X","Linw Y","Min","Measured","Max","Units","Result"])
                self.pushButton_Measure.setDisabled(True)
                QMessageBox.warning(self,"WARNING!!!","Invalid File Passed")
    ##########################################################################################################
    def SaveReport(self):
        workbook = load_workbook(filename="Reports/BrowseTest/BrowseTestTemplate.xlsx")
        # open workbook
        sheet = workbook.active
        sheet["A5"] = self.inputsheet["A1"].value
        sheet["A6"] = self.inputsheet["A2"].value
        sheet["A7"] = self.inputsheet["A3"].value
        sheet["A8"] = self.inputsheet["A4"].value
        sheet["A9"] = self.inputsheet["A5"].value
        sheet["C3"] = datetime.datetime.now().strftime("%d/%m/%Y")
        sheet["D3"] = datetime.datetime.now().strftime("%I:%M%p")

        for i in range(0, self.tableWidget.rowCount()):
            sheet[f'''A{i + 12}'''] = self.tableWidget.item(i, 0).text()
            sheet[f'''B{i + 12}'''] = self.tableWidget.item(i, 1).text()
            sheet[f'''C{i + 12}'''] = self.inputsheet[f'''B{i + 8}'''].value
            sheet[f'''D{i + 12}'''] = self.tableWidget.item(i, 2).text()
            sheet[f'''E{i + 12}'''] = self.inputsheet[f'''D{i + 8}'''].value
            sheet[f'''F{i + 12}'''] = self.tableWidget.item(i, 3).text()
            sheet[f'''G{i + 12}'''] = self.tableWidget.item(i, 4).text()
            sheet[f'''H{i + 12}'''] = self.tableWidget.item(i, 5).text()
          #  sheet[f'''H{i + 11}'''] = self.tableWidget.item(i, 6).text()
            sheet[f'''I{i + 12}'''] = self.tableWidget.item(i, 7).text()

        self.set_border(sheet, f'''A11:I{12 + i}''')

        # save the file
        outfile = "Reports/BrowseTest/BrowseTest" + datetime.datetime.now().strftime('%d_%m_%Y_%H_%M_%S') + '.xlsx'
        workbook.save(filename=outfile)
        QMessageBox.information(self, "Browse Test", "Reports Saved to " + outfile)

    #################################################################################################################
    def set_border(self, worksheet, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in worksheet[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)