import datetime
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

class Get_Reports():
    def set_border(self,worksheet, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in worksheet[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def Generate_Report_SelfTest(self,table_widget):
        # load excel file
        workbook = load_workbook(filename="Reports/SelfTest/SelfTestTemplate.xlsx")
        # open workbook
        sheet = workbook.active

        for i in range(0,table_widget.rowCount()):
            sheet[f'''A{i+8}'''] = table_widget.item(i,0).text()
            sheet[f'''B{i + 8}'''] = table_widget.item(i, 1).text()
            sheet[f'''C{i + 8}'''] = table_widget.item(i, 2).text()
            sheet[f'''D{i + 8}'''] = table_widget.item(i, 3).text()
            sheet[f'''E{i + 8}'''] = table_widget.item(i, 4).text()
            sheet[f'''F{i + 8}'''] = table_widget.item(i, 5).text()
            sheet[f'''G{i + 8}'''] = table_widget.item(i, 6).text()
            sheet[f'''H{i + 8}'''] = table_widget.item(i, 7).text()

        self.set_border(sheet, f'''A8:H{8+i}''')

        # save the file
        workbook.save(filename="Reports/SelfTest/SelfTest"+datetime.datetime.now().strftime('%d_%m_%Y_%H_%M_%S')+'.xlsx')

    def Generate_Report_PTPMan(self, table_widget):
        # load excel file
        workbook = load_workbook(filename="Reports/PTPManual/PTPManualTemplate.xlsx")
        # open workbook
        sheet = workbook.active

        for i in range(0, table_widget.rowCount()):
            sheet[f'''A{i + 8}'''] = table_widget.item(i, 0).text()
            sheet[f'''B{i + 8}'''] = table_widget.item(i, 1).text()
            sheet[f'''C{i + 8}'''] = table_widget.item(i, 2).text()
            sheet[f'''D{i + 8}'''] = table_widget.item(i, 3).text()
            sheet[f'''E{i + 8}'''] = table_widget.item(i, 4).text()
            sheet[f'''F{i + 8}'''] = table_widget.item(i, 5).text()
            sheet[f'''G{i + 8}'''] = table_widget.item(i, 6).text()
            sheet[f'''H{i + 8}'''] = table_widget.item(i, 7).text()

        self.set_border(sheet, f'''A8:H{8 + i}''')

        # save the file
        workbook.save(
            filename="Reports/PTPManual/PTPManualTest" + datetime.datetime.now().strftime('%d_%m_%Y_%H_%M_%S') + '.xlsx')

    def Generate_Report_PTPAut(self, table_widget):
        # load excel file
        workbook = load_workbook(filename="Reports/PTPAuto/PTPAutoTemplate.xlsx")
        # open workbook
        sheet = workbook.active

        for i in range(0, table_widget.rowCount()):
            sheet[f'''A{i + 8}'''] = table_widget.item(i, 0).text()
            sheet[f'''B{i + 8}'''] = table_widget.item(i, 1).text()
            sheet[f'''C{i + 8}'''] = table_widget.item(i, 2).text()
            sheet[f'''D{i + 8}'''] = table_widget.item(i, 3).text()
            sheet[f'''E{i + 8}'''] = table_widget.item(i, 4).text()
            sheet[f'''F{i + 8}'''] = table_widget.item(i, 5).text()
            sheet[f'''G{i + 8}'''] = table_widget.item(i, 6).text()
            sheet[f'''H{i + 8}'''] = table_widget.item(i, 7).text()

        self.set_border(sheet, f'''A8:H{8 + i}''')

        # save the file
        workbook.save(
            filename="Reports/PTPAuto/PTPAutoTest" + datetime.datetime.now().strftime('%d_%m_%Y_%H_%M_%S') + '.xlsx')