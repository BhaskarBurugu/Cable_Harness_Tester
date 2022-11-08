# Writing to an excel
# sheet using Python
'''
import xlwt
from xlwt import Workbook

# Workbook is created
wb = Workbook()

# add_sheet is used to create sheet.
sheet1 = wb.add_sheet('Sheet 1')

sheet1.write(1, 0, 'ISBT DEHRADUN')
sheet1.write(2, 0, 'SHASTRADHARA')
sheet1.write(3, 0, 'CLEMEN TOWN')
sheet1.write(4, 0, 'RAJPUR ROAD')
sheet1.write(5, 0, 'CLOCK TOWER')
sheet1.write(0, 1, 'ISBT DEHRADUN')
sheet1.write(0, 2, 'SHASTRADHARA')
sheet1.write(0, 3, 'CLEMEN TOWN')
sheet1.write(0, 4, 'RAJPUR ROAD')
sheet1.write(0, 5, 'CLOCK TOWER')

wb.save('xlwt example.xls')
'''
from openpyxl import load_workbook

'''
import datetime

from openpyxl import load_workbook

# load excel file
from openpyxl.styles import Border, Side

workbook = load_workbook(filename="Reports/SelfTest/SelfTestTemplate.xlsx")

# open workbook
sheet = workbook.worksheets[0]
i = 0
index = f'D{i+8}'
# modify the desired cell
sheet[f'A{i+8}'] = "Burug"

def set_border(ws, cell_range):
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

    rows = ws.iter_rows(cell_range)
    print(rows)
    for row in rows:
        for cell in row:
            cell.border = border

set_border(sheet, 'A8:H20')

# save the file
workbook.save(filename="output.xlsx")

print(datetime.datetime.now().strftime('%d_%m_%Y_%H_%M_%S'))
'''

from openpyxl.styles import Border

wb = load_workbook(filename = "Reports/SelfTest/SelfTestTemplate.xlsx")
ws = wb.worksheets[0]
from openpyxl.styles import Border, Side

def set_border1(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def set_border(ws, cell_range):
    rows = ws[cell_range]
    for row in rows:
        if row == rows[0][0] or row == rows[0][-1] or row == rows[-1][0] or row == rows[-1][-1]:
            pass
        else:
            row[0].border = Border(left=Side(style='thin'))
            row[-1].border = Border(right=Side(style='thin'))
        for c in rows[0]:
            c.border = Border(top=Side(style='thin'))
        for c in rows[-1]:
            c.border = Border(bottom=Side(style='thin'))
    rows[0][0].border = Border(left=Side(style='thin'), top=Side(style='thin'))
    rows[0][-1].border = Border(right=Side(style='thin'), top=Side(style='thin'))
    rows[-1][0].border = Border(left=Side(style='thin'), bottom=Side(style='thin'))
    rows[-1][-1].border = Border(right=Side(style='thin'), bottom=Side(style='thin'))

set_border(ws, 'A8:H120')
wb.save(filename="output.xlsx")
