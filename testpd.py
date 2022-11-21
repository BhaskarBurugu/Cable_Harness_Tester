
'''
from openpyxl import load_workbook
frompinlist = []
topinlist = []
minvallist = []
maxvallist = []
workbook = load_workbook(filename="Reports/SelfTest/SelfTest10_11_2022_16_22_29.xlsx")
sheet = workbook.active
nr = sheet.max_row
for i in range(8, nr):
    fp=sheet.cell(i, 2).value
    frompinlist.append(fp)
    topinlist.append(sheet.cell(i,3).value)
    minvallist.append(sheet.cell(i,6).value)
    maxvallist.append(sheet.cell(i, 7).value)

    print(fp[4:])
'''
#frominlist_byte = bytearray(frompinlist)
#topinlist_byte =  bytearray(topinlist)
#print((10).to_bytes(1, byteorder='big'))

def Resistance_Format(resistance = 100e3):
    if (resistance >= 0.0 and resistance <1000):
        return f'''{resistance:.2f} \u03A9'''
    elif (resistance >= 1e3 and resistance < 1000e3):
        return f'''{(resistance/1e3):.2f} K\u03A9'''
    elif resistance >= 1e6 :
        return f'''{(resistance/1e6):.2f} M\u03A9'''


print(Resistance_Format(float('nan')))

